# Marks Dvojeglazovs
# Riga Technical University, 2023/2024

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import openpyxl
import os
import ui

header = ["Numurs", "Datums", "url",
          "Marka", "Motors", "Nobraukums", "Virsbūves tips", "Izlaiduma gads", "Ātr.kārba", "Krāsa", "Tehniskā apskate", "Cena", "Vieta",
          "Stūres hidropastiprinātājs", "Elektr. stūres pastiprinātājs", "Kondicionieris", "Klimata kontrole", "Autonomais sildītājs", "Salona gaisa filtrs", "Borta dators", "Riepu spiediena kontrole", "Parkošanās sens. aizmugurē", "Parkošanās sens. priekšā", "Atpakaļskata kamera", "Nakts redzamības kamera", "Panorāmas redzam. kameras", "Distances kontrole", "Lietus sensors", "Kruīza kontrole", "Adaptīvā kruīzkontrole", "Gājēju detektors", "Automātiska parkošana", "Ceļa zīmju atpazīš. sistēma", "Klusuma zonu asistents", "Palīgsistēma braukš. joslās", "Avārijas bremzēšanas sist.", "Keyless Go", "Sistēma Starts-Stop", "El. bagāžnieka aizvēršana", "El. aizmugures saulessargs", "El. durvju aizvēršana", "Jumta reliņi", "Sakabes āķis", "Lūka", "Panorāmas lūka", "Pilnpiedziņa 4x4", "Pneimopiekare", "Spoileris", "Sliekšņi", "Sporta pakete", "Servisa grāmatiņa", "Vieglmetāla diski",
          "Ādas apdare", "Roku balsti", "Tonēti aizmugurējie logi", "Saulessargi logiem", "Isofix stiprinājumi", "Ledusskapis",
          "Regulējama", "El. regulējama", "Daudzfunkcionāla", "Sporta", "Apsildāma", "Ar atmiņu",
          "El. regulējami", "Apsildāmi", "Sporta", "Recaro", "Ventilējami", "Masāžas", "Ar atmiņu",
          "Xenona", "Bi xenona", "LED", "LED bremžugunis", "Papild. bremžu signāls", "Miglas lukturi", "Lampu mazgātāji", "Automāt. tuvās gaismas", "Automāt. tālās gaismas", "Adaptīvās tālās gaismas",
          "El. regulējami", "Apsildāmi", "Aptumšojošie", "Sporta", "El. nolokāmi", "Ar atmiņu",
          "ABS", "Centrālā atslēga", "Signalizācija", "Imobilaizers", "Air-bag", "ESP", "ASR", "Marķējums",
          "FM/AM", "CD", "CD mainītājs", "DVD", "DVD mainītājs", "MP3", "USB", "SDcard", "HDD", "TV", "LCD", "Navigācija", "Tel./mob.", "Bluetooth", "Hands-free", "Subwoofer"]


def generate_file(chosen_mark, mode):
    if mode == "Real-time":
        if not os.path.exists("wb"):
            os.mkdir("wb")
        get_urls(chosen_mark)
        get_info(chosen_mark)
        ui.show_data(chosen_mark)
    elif mode == "Test":
        print("Using test data from /test-wb/ folder")


def get_urls(chosen_mark):
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    wb = openpyxl.Workbook()
    wb.save("wb/" + chosen_mark + ".xlsx")
    wb = openpyxl.load_workbook("wb/" + chosen_mark + ".xlsx")
    ws = wb.active
    for j in range(len(header)):
        ws.cell(row=1, column=j + 1).value = header[j]
    wb.save("wb/" + chosen_mark + ".xlsx")

    url = "https://www.ss.com/lv/transport/cars/" + chosen_mark + "/sell/"
    driver.get(url)

    pages = driver.find_elements(By.CLASS_NAME, "navi")
    if pages:
        pages = pages[0].get_attribute("href").split("page")[1].split(".")[0]
    else:
        pages = 1

    for j in range(int(pages)):
        if j != 0:
            driver.get(url + "page" + str(j + 1) + ".html")
        cars = driver.find_elements(By.CLASS_NAME, "am")
        for k in range(len(cars)):
            ws.cell(row=k + 2 + j * 30, column=1).value = k + 1 + j * 30
            ws.cell(row=k + 2 + j * 30, column=3).value = cars[k].get_attribute("href")
    wb.save("wb/" + chosen_mark + ".xlsx")

def get_info(chosen_mark):
    service = Service()
    option = webdriver.ChromeOptions()
    driver = webdriver.Chrome(service=service, options=option)

    wb = openpyxl.load_workbook("wb/" + chosen_mark + ".xlsx")
    ws = wb.active

    for i in range (2, ws.max_row + 1):
        driver.get(ws.cell(row=i, column=3).value)
        ws.cell(row=i, column=2).value = driver.find_element(By.XPATH, ".//*").text.split("Datums: ")[1].split("Pārsūtīt")[0]
        if driver.find_element(By.XPATH, ".//*").text.split("Vieta: ")[1].split("Pievienot")[0].__contains__("Adrese"):
            ws.cell(row=i, column=13).value = driver.find_element(By.XPATH, ".//*").text.split("Vieta: ")[1].split("Adrese:")[0]
        else:
            ws.cell(row=i, column=13).value = driver.find_element(By.XPATH, ".//*").text.split("Vieta: ")[1].split("Pievienot")[0]

        try:
            ws.cell(row=i, column=4).value = driver.find_element(By.ID, "tdo_31").text
        except:
            pass
        try:
            ws.cell(row=i, column=5).value = driver.find_element(By.ID, "tdo_15").text
        except:
            pass
        try:
            ws.cell(row=i, column=5).value = driver.find_element(By.ID, "tdo_34").text
        except:
            pass
        try:
            ws.cell(row=i, column=6).value = driver.find_element(By.ID, "tdo_16").text.replace(" ", "")
        except:
            pass
        try:
            ws.cell(row=i, column=7).value = driver.find_element(By.ID, "tdo_32").text
        except:
            pass
        try:
            ws.cell(row=i, column=8).value = driver.find_element(By.ID, "tdo_18").text.split(" ")[0]
        except:
            pass
        try:
            ws.cell(row=i, column=9).value = driver.find_element(By.ID, "tdo_35").text
        except:
            pass
        try:
            ws.cell(row=i, column=10).value = driver.find_element(By.ID, "tdo_17").text.split(" ")[0]
        except:
            pass
        try:
            ws.cell(row=i, column=11).value = driver.find_element(By.ID, "tdo_223").text
        except:
            pass
        try:
            ws.cell(row=i, column=12).value = driver.find_element(By.ID, "tdo_8").text.replace(" €", "").replace(" ", "")
        except:
            pass

        info = driver.find_elements(By.CLASS_NAME, "auto_c")
        for j in range(len(info)):
            info[j] = info[j].text
        print(info)
        for j in range(len(info)):
            for k in range(len(header)):
                if header[k] == info[j]:
                    ws.cell(row=i, column=k + 1).value = "true"
                    break

        for j in range(1, 13):
            if ws.cell(row=i, column=j).value is None:
                ws.cell(row=i, column=j).value = "N/A"
        for j in range(14, ws.max_column + 1):
            if ws.cell(row=i, column=j).value is None:
                ws.cell(row=i, column=j).value = "false"

        wb.save("wb/" + chosen_mark + ".xlsx")